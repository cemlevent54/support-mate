import json
import base64
import csv
import logging
import os
import math
from datetime import datetime
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, KeepTogether, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.graphics.shapes import Drawing, String
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.lineplots import LinePlot
from reportlab.graphics.widgets.markers import makeMarker
from io import BytesIO
from jinja2 import Environment, FileSystemLoader
from xhtml2pdf import pisa
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

logger = logging.getLogger(__name__)

def create_status_pie_chart(status_data, language='tr'):
    """Status dagilimi icin pie chart olustur"""
    if not status_data:
        return Spacer(1, 10)
    
    # Renk tanimlari
    color_map = {
        'PENDING': colors.red,
        'IN_PROGRESS': colors.blue,
        'DONE': colors.green,
        'OPEN': colors.red,
        'IN_REVIEW': colors.orange,
        'CLOSED': colors.green,
        'default': colors.grey
    }
    
    # Veriyi hazirla
    data = []
    total = sum(status_data.values())
    
    for status, count in status_data.items():
        if count > 0:
            percentage = (count / total) * 100 if total > 0 else 0
            color = color_map.get(status, color_map['default'])
            data.append({
                'status': status,
                'count': count,
                'percentage': percentage,
                'color': color
            })
    
    if not data:
       return [Spacer(1, 10)]
    
    # 3D Exploded Pie chart olustur - sol tarafa
    drawing = Drawing(500, 250)
    pie = Pie()
    pie.x = 160  # Ortali konum (500-180)/2 = 160
    pie.y = 80  # Basliktan uzaklastir
    pie.width = 180
    pie.height = 180
    pie.slices.strokeWidth = 2
    pie.slices.strokeColor = colors.white
    pie.slices.popout = 10  # Exploded effect
    pie.slices.fontSize = 8
    pie.slices.fontName = 'Helvetica-Bold'
    pie.slices.fontColor = colors.black
    
    # Verileri pie chart'a ekle
    pie.data = [item['count'] for item in data]
    pie.labels = [f"{item['status']} {item['count']} , %{item['percentage']:.1f}" for item in data]  # Tek satirda tum bilgiler
    pie.slices.strokeWidth = 2
    pie.slices.strokeColor = colors.white
    
    # Renkleri ayarla
    for i, item in enumerate(data):
        pie.slices[i].fillColor = item['color']
    
    drawing.add(pie)
    
    # Aciklayici metin olustur
    legend_text = []
    for item in data:
        status_name = item['status']
        if language == 'tr':
            if status_name == 'PENDING':
                status_name = 'Bekleyen'
            elif status_name == 'IN_PROGRESS':
                status_name = 'Devam Eden'
            elif status_name == 'DONE':
                status_name = 'Tamamlanan'
            elif status_name == 'OPEN':
                status_name = 'Acik'
            elif status_name == 'IN_REVIEW':
                status_name = 'Incelemede'
            elif status_name == 'CLOSED':
                status_name = 'Kapali'
        
        # Renk kutusu ile birlikte legend
        legend_text.append(f"■ {status_name}")
        legend_text.append(f"   {item['count']} ,")
        legend_text.append(f"   %{item['percentage']:.1f}")
        legend_text.append("")
    
    # Baslik
    title_text = "Status Dagilimi" if language == 'tr' else "Status Distribution"
    title_style = ParagraphStyle(
        'ChartTitle',
        parent=getSampleStyleSheet()['Heading3'],
        fontSize=14,
        textColor=colors.black,
        alignment=1,  # Ortali
        spaceBefore=10,
        spaceAfter=20  # Daha fazla bosluk
    )
    
    # Aciklayici metin - pie chart'in yanina
    legend_style = ParagraphStyle(
        'Legend',
        parent=getSampleStyleSheet()['Normal'],
        fontSize=10,
        spaceAfter=6,
        leftIndent=240,  # Pie chart'in yanina
        textColor=colors.HexColor('#424242'),
        fontName='Helvetica'
    )
    
    elements = []
    elements.append(Paragraph(title_text, title_style))
    
    # Toplam sayi bilgisi - pie chart'in altina
    total_text = f"Toplam: {total}" if language == 'tr' else f"Total: {total}"
    total_style = ParagraphStyle(
        'Total',
        parent=getSampleStyleSheet()['Normal'],
        fontSize=14,
        spaceAfter=15,
        textColor=colors.HexColor('#1976d2'),
        alignment=1,  # Ortali
        fontName='Helvetica-Bold'
    )
    
    elements.append(drawing)
    elements.append(Paragraph(total_text, total_style))
    
    return elements

def create_priority_pie_chart(priority_data, language='tr'):
    """Priority dagilimi icin pie chart olustur"""
    if not priority_data:
        return []
    
    # Toplam sayiyi hesapla
    total = sum(priority_data.values())
    if total == 0:
        return []
    
    # Verileri hazirla
    data = []
    for priority, count in priority_data.items():
        if count > 0:
            percentage = (count / total) * 100
            
            # Priority isimlerini cevir
            priority_name = priority
            if language == 'tr':
                if priority == 'LOW':
                    priority_name = 'Dusuk'
                elif priority == 'MEDIUM':
                    priority_name = 'Orta'
                elif priority == 'HIGH':
                    priority_name = 'Yuksek'
                elif priority == 'CRITICAL':
                    priority_name = 'Kritik'
            
            # Renk haritasi
            color_map = {
                'LOW': colors.HexColor('#4CAF50'),      # Yesil
                'MEDIUM': colors.HexColor('#2196F3'),   # Mavi
                'HIGH': colors.HexColor('#FF9800'),     # Turuncu
                'CRITICAL': colors.HexColor('#F44336')  # Kirmizi
            }
            
            data.append({
                'priority': priority_name,
                'count': count,
                'percentage': percentage,
                'color': color_map.get(priority, colors.HexColor('#9E9E9E'))
            })
    
    if not data:
        return []
    
    # 3D Exploded Pie chart olustur - sol tarafa
    drawing = Drawing(500, 250)
    pie = Pie()
    pie.x = 160  # Ortali konum
    pie.y = 80  # Basliktan uzaklastir
    pie.width = 180
    pie.height = 180
    pie.slices.strokeWidth = 2
    pie.slices.strokeColor = colors.white
    pie.slices.popout = 10  # Exploded effect
    pie.slices.fontSize = 8
    pie.slices.fontName = 'Helvetica-Bold'
    pie.slices.fontColor = colors.black
    
    # Verileri pie chart'a ekle
    pie.data = [item['count'] for item in data]
    pie.labels = [f"{item['priority']} {item['count']} , %{item['percentage']:.1f}" for item in data]  # Tek satirda tum bilgiler
    pie.slices.strokeWidth = 2
    pie.slices.strokeColor = colors.white
    
    # Renkleri ayarla
    for i, item in enumerate(data):
        pie.slices[i].fillColor = item['color']
    
    drawing.add(pie)
    
    # Baslik
    title_text = "Oncelik Dagilimi" if language == 'tr' else "Priority Distribution"
    title_style = ParagraphStyle(
        'ChartTitle',
        parent=getSampleStyleSheet()['Heading3'],
        fontSize=14,
        textColor=colors.black,
        alignment=1,  # Ortali
        spaceBefore=10,
        spaceAfter=20  # Daha fazla bosluk
    )
    
    elements = []
    elements.append(Paragraph(title_text, title_style))
    
    # Toplam sayi bilgisi - pie chart'in altina
    total_text = f"Toplam: {total}" if language == 'tr' else f"Total: {total}"
    total_style = ParagraphStyle(
        'Total',
        parent=getSampleStyleSheet()['Normal'],
        fontSize=14,
        spaceAfter=15,
        textColor=colors.HexColor('#1976d2'),
        alignment=1,  # Ortali
        fontName='Helvetica-Bold'
    )
    
    elements.append(drawing)
    elements.append(Paragraph(total_text, total_style))
    
    return elements

def create_resolution_pie_chart(resolution_data, language='tr'):
    """Resolution sureleri icin pie chart olustur"""
    if not resolution_data:
        return []
    
    # Toplam sayiyi hesapla
    total = sum(resolution_data.values())
    if total == 0:
        return []
    
    # Verileri hazirla
    data = []
    for resolution, count in resolution_data.items():
        if count > 0:
            percentage = (count / total) * 100
            
            # Resolution isimlerini cevir
            resolution_name = resolution
            if language == 'tr':
                if resolution == 'inaday':
                    resolution_name = 'Bir gunde'
                elif resolution == 'inaweek':
                    resolution_name = 'Bir haftada'
                elif resolution == 'inamonth':
                    resolution_name = 'Bir ayda'
                elif resolution == 'notresolved':
                    resolution_name = 'Cozulmedi'
            else:
                # Ingilizce icin de cevir
                if resolution == 'inaday':
                    resolution_name = 'In a day'
                elif resolution == 'inaweek':
                    resolution_name = 'In a week'
                elif resolution == 'inamonth':
                    resolution_name = 'In a month'
                elif resolution == 'notresolved':
                    resolution_name = 'Not resolved'
            
            # Renk haritasi
            color_map = {
                'inaday': colors.HexColor('#4CAF50'),      # Yesil
                'inaweek': colors.HexColor('#2196F3'),     # Mavi
                'inamonth': colors.HexColor('#FF9800'),    # Turuncu
                'notresolved': colors.HexColor('#F44336')  # Kirmizi
            }
            
            data.append({
                'resolution': resolution_name,
                'count': count,
                'percentage': percentage,
                'color': color_map.get(resolution, colors.HexColor('#9E9E9E'))
            })
    
    if not data:
        return []
    
    # 3D Exploded Pie chart olustur - sol tarafa
    drawing = Drawing(500, 250)
    pie = Pie()
    pie.x = 160  # Ortali konum
    pie.y = 80  # Basliktan uzaklastir
    pie.width = 180
    pie.height = 180
    pie.slices.strokeWidth = 2
    pie.slices.strokeColor = colors.white
    pie.slices.popout = 10  # Exploded effect
    pie.slices.fontSize = 8
    pie.slices.fontName = 'Helvetica-Bold'
    pie.slices.fontColor = colors.black
    
    # Verileri pie chart'a ekle
    pie.data = [item['count'] for item in data]
    pie.labels = [f"{item['resolution']} {item['count']} , %{item['percentage']:.1f}" for item in data]  # Tek satirda tum bilgiler
    pie.slices.strokeWidth = 2
    pie.slices.strokeColor = colors.white
    
    # Renkleri ayarla
    for i, item in enumerate(data):
        pie.slices[i].fillColor = item['color']
    
    drawing.add(pie)
    
    # Baslik
    title_text = "Cozum Sureleri" if language == 'tr' else "Resolution Times"
    title_style = ParagraphStyle(
        'ChartTitle',
        parent=getSampleStyleSheet()['Heading3'],
        fontSize=14,
        textColor=colors.black,
        alignment=1,  # Ortali
        spaceBefore=10,
        spaceAfter=20  # Daha fazla bosluk
    )
    
    elements = []
    elements.append(Paragraph(title_text, title_style))
    
    # Toplam sayi bilgisi - pie chart'in altina
    total_text = f"Toplam: {total}" if language == 'tr' else f"Total: {total}"
    total_style = ParagraphStyle(
        'Total',
        parent=getSampleStyleSheet()['Normal'],
        fontSize=14,
        spaceAfter=15,
        textColor=colors.HexColor('#1976d2'),
        alignment=1,  # Ortali
        fontName='Helvetica-Bold'
    )
    
    elements.append(drawing)
    elements.append(Paragraph(total_text, total_style))
    
    return elements

def create_user_verification_pie_chart(users_data, language='tr'):
    """User verification durumu icin pie chart olustur"""
    if not users_data or users_data.get('total') is None:
        return []
    
    total = users_data.get('total', 0)
    verified = users_data.get('verifiedUsers', 0)
    blocked = total - verified  # Manuel hesaplama
    
    if total == 0:
        return []
    
    # Verileri hazirla
    data = []
    
    if verified > 0:
        verified_percentage = (verified / total) * 100
        data.append({
            'status': 'Dogrulanmis' if language == 'tr' else 'Verified',
            'count': verified,
            'percentage': verified_percentage,
            'color': colors.HexColor('#4CAF50')  # Yesil
        })
    
    if blocked > 0:
        blocked_percentage = (blocked / total) * 100
        data.append({
            'status': 'Engellenmis' if language == 'tr' else 'Blocked',
            'count': blocked,
            'percentage': blocked_percentage,
            'color': colors.HexColor('#F44336')  # Kirmizi
        })
    
    if not data:
        return []
    
    # 3D Exploded Pie chart olustur
    drawing = Drawing(500, 250)
    pie = Pie()
    pie.x = 160  # Ortali konum
    pie.y = 80  # Basliktan uzaklastir
    pie.width = 180
    pie.height = 180
    pie.slices.strokeWidth = 2
    pie.slices.strokeColor = colors.white
    pie.slices.popout = 10  # Exploded effect
    pie.slices.fontSize = 8
    pie.slices.fontName = 'Helvetica-Bold'
    pie.slices.fontColor = colors.black
    
    # Verileri pie chart'a ekle
    pie.data = [item['count'] for item in data]
    pie.labels = [f"{item['status']} {item['count']} , %{item['percentage']:.1f}" for item in data]  # Tek satirda tum bilgiler
    pie.slices.strokeWidth = 2
    pie.slices.strokeColor = colors.white
    
    # Renkleri ayarla
    for i, item in enumerate(data):
        pie.slices[i].fillColor = item['color']
    
    drawing.add(pie)
    
    # Baslik
    title_text = "Kullanici Dogrulama Durumu" if language == 'tr' else "User Verification Status"
    title_style = ParagraphStyle(
        'ChartTitle',
        parent=getSampleStyleSheet()['Heading3'],
        fontSize=14,
        textColor=colors.black,
        alignment=1,  # Ortali
        spaceBefore=10,
        spaceAfter=50  # Cok daha fazla bosluk
    )
    
    elements = []
    elements.append(Paragraph(title_text, title_style))
    
    # Pie chart'i basliktan sonra ekle
    elements.append(drawing)
    
    # Toplam sayi bilgisi - pie chart'in altina
    total_text = f"Toplam: {total}" if language == 'tr' else f"Total: {total}"
    total_style = ParagraphStyle(
        'Total',
        parent=getSampleStyleSheet()['Normal'],
        fontSize=14,
        spaceAfter=15,
        textColor=colors.HexColor('#1976d2'),
        alignment=1,  # Ortali
        fontName='Helvetica-Bold'
    )
    
    elements.append(Paragraph(total_text, total_style))
    
    return elements

def create_user_role_pie_chart(roles_data, language='tr'):
    """User role dagilimi icin pie chart olustur"""
    if not roles_data:
        return []
    
    # Toplam sayiyi hesapla
    total = sum(r.get('count', r.get('numberOfUsers', 0)) for r in roles_data)
    if total == 0:
        return []
    
    # Verileri hazirla
    data = []
    for role in roles_data:
        count = role.get('count', role.get('numberOfUsers', 0))
        if count > 0:
            percentage = (count / total) * 100
            role_name = role.get('roleName', 'Unknown')
            
            # Role isimlerini cevir
            if language == 'tr':
                if role_name == 'ADMIN':
                    role_name = 'Admin'
                elif role_name == 'CUSTOMER_SUPPORTER':
                    role_name = 'Musteri Destek'
                elif role_name == 'EMPLOYEE':
                    role_name = 'Calisan'
                elif role_name == 'LEADER':
                    role_name = 'Lider'
                elif role_name == 'USER':
                    role_name = 'Kullanici'
            
            # Renk haritasi - sadece sistemdeki 5 rol
            color_map = {
                'ADMIN': colors.HexColor('#F44336'),           # Kirmizi
                'CUSTOMER_SUPPORTER': colors.HexColor('#4CAF50'),  # Yesil
                'EMPLOYEE': colors.HexColor('#FFEB3B'),        # Sari
                'LEADER': colors.HexColor('#E91E63'),          # Pembe
                'USER': colors.HexColor('#3F51B5')             # Lacivert
            }
            
            # Role name'e gore renk bul
            role_key = role.get('roleName', 'UNKNOWN').upper()
            selected_color = color_map.get(role_key, colors.HexColor('#9E9E9E'))
            
            data.append({
                'role': role_name,
                'count': count,
                'percentage': percentage,
                'color': selected_color
            })
    
    if not data:
        return []
    
    # 3D Exploded Pie chart olustur
    drawing = Drawing(500, 250)
    pie = Pie()
    pie.x = 160  # Ortali konum
    pie.y = 80  # Basliktan uzaklastir
    pie.width = 180
    pie.height = 180
    pie.slices.strokeWidth = 2
    pie.slices.strokeColor = colors.white
    pie.slices.popout = 10  # Exploded effect
    pie.slices.fontSize = 8
    pie.slices.fontName = 'Helvetica-Bold'
    pie.slices.fontColor = colors.black
    
    # Verileri pie chart'a ekle
    pie.data = [item['count'] for item in data]
    pie.labels = [f"{item['role']} {item['count']} , %{item['percentage']:.1f}" for item in data]  # Tek satirda tum bilgiler
    pie.slices.strokeWidth = 2
    pie.slices.strokeColor = colors.white
    
    # Renkleri ayarla
    for i, item in enumerate(data):
        pie.slices[i].fillColor = item['color']
    
    drawing.add(pie)
    
    # Baslik
    title_text = "Rol Bazli Kullanici Dagilimi" if language == 'tr' else "User Distribution by Role"
    title_style = ParagraphStyle(
        'ChartTitle',
        parent=getSampleStyleSheet()['Heading3'],
        fontSize=14,
        textColor=colors.black,
        alignment=1,  # Ortali
        spaceBefore=10,
        spaceAfter=50  # Cok daha fazla bosluk
    )
    
    elements = []
    elements.append(Paragraph(title_text, title_style))
    
    # Pie chart'i basliktan sonra ekle
    elements.append(drawing)
    
    # Toplam sayi bilgisi - pie chart'in altina
    total_text = f"Toplam: {total}" if language == 'tr' else f"Total: {total}"
    total_style = ParagraphStyle(
        'Total',
        parent=getSampleStyleSheet()['Normal'],
        fontSize=14,
        spaceAfter=15,
        textColor=colors.HexColor('#1976d2'),
        alignment=1,  # Ortali
        fontName='Helvetica-Bold'
    )
    
    elements.append(Paragraph(total_text, total_style))
    
    return elements

def create_daily_ticket_line_plot(tickets_data, language='tr'):
    """Daily ticket distribution için line plot oluştur"""
    if not tickets_data or not tickets_data.get('dates'):
        return []
    
    dates_data = tickets_data['dates']
    if not dates_data:
        return []
    
    # X ekseni için günleri hazırla (1'den başlayarak)
    x_values = list(range(1, len(dates_data) + 1))
    
    # Y ekseni için verileri hazırla
    total_data = []
    open_data = []
    review_data = []
    progress_data = []
    closed_data = []
    
    for i, date_info in enumerate(dates_data):
        x_val = i + 1
        
        # Total tickets
        total_count = date_info.get('totalTickets', 0)
        total_data.append((x_val, total_count))
        
        # Open tickets
        open_count = date_info.get('openTickets', 0)
        open_data.append((x_val, open_count))
        
        # Review tickets
        review_count = date_info.get('reviewTickets', 0)
        review_data.append((x_val, review_count))
        
        # Progress tickets
        progress_count = date_info.get('progressTickets', 0)
        progress_data.append((x_val, progress_count))
        
        # Closed tickets
        closed_count = date_info.get('closedTickets', 0)
        closed_data.append((x_val, closed_count))
    
    # Line plot oluştur
    drawing = Drawing(500, 300)
    
    lp = LinePlot()
    lp.x = 80
    lp.y = 100  # Y pozisyonunu artır
    lp.height = 250  # Yüksekliği artır
    lp.width = 350
    lp.data = [total_data, open_data, review_data, progress_data, closed_data]
    lp.joinedLines = 1
    
    # Her çizgi için farklı marker ve renk
    lp.lines[0].symbol = makeMarker('FilledCircle')  # Total - Mavi
    lp.lines[0].strokeColor = colors.blue
    lp.lines[0].strokeWidth = 2
    
    lp.lines[1].symbol = makeMarker('Circle')  # Open - Kırmızı
    lp.lines[1].strokeColor = colors.red
    lp.lines[1].strokeWidth = 2
    
    lp.lines[2].symbol = makeMarker('Diamond')  # Review - Turuncu
    lp.lines[2].strokeColor = colors.orange
    lp.lines[2].strokeWidth = 2
    
    lp.lines[3].symbol = makeMarker('Square')  # Progress - Mavi
    lp.lines[3].strokeColor = colors.HexColor('#2196F3')
    lp.lines[3].strokeWidth = 2
    
    lp.lines[4].symbol = makeMarker('Triangle')  # Closed - Yeşil
    lp.lines[4].strokeColor = colors.green
    lp.lines[4].strokeWidth = 2
    
    lp.lineLabelFormat = '%2.0f'
    lp.strokeColor = colors.black
    
    # X ekseni ayarları
    lp.xValueAxis.valueMin = 0
    lp.xValueAxis.valueMax = len(dates_data) + 1
    lp.xValueAxis.valueSteps = x_values
    lp.xValueAxis.labelTextFormat = '%d'
    
    # Y ekseni ayarları - Dinamik değer (en yüksek + 10)
    max_value = max([
        max([y for x, y in total_data]) if total_data else 0,
        max([y for x, y in open_data]) if open_data else 0,
        max([y for x, y in review_data]) if review_data else 0,
        max([y for x, y in progress_data]) if progress_data else 0,
        max([y for x, y in closed_data]) if closed_data else 0
    ])
    
    lp.yValueAxis.valueMin = 0
    lp.yValueAxis.valueMax = max_value + 10  # En yüksek değer + 10
    lp.yValueAxis.valueSteps = list(range(0, max_value + 11, max(1, (max_value + 10) // 5)))
    
    drawing.add(lp)
    
    # Başlık
    title_text = "Günlük Ticket Dagilimi" if language == 'tr' else "Daily Ticket Distribution"
    title_style = ParagraphStyle(
        'ChartTitle',
        parent=getSampleStyleSheet()['Heading3'],
        fontSize=14,
        textColor=colors.black,
        alignment=1,  # Ortalı
        spaceBefore=10,
        spaceAfter=80  # Daha fazla boşluk
    )
    
    # Toplam değerleri hesapla
    total_sum = sum([y for x, y in total_data]) if total_data else 0
    open_sum = sum([y for x, y in open_data]) if open_data else 0
    review_sum = sum([y for x, y in review_data]) if review_data else 0
    progress_sum = sum([y for x, y in progress_data]) if progress_data else 0
    closed_sum = sum([y for x, y in closed_data]) if closed_data else 0
    
    # Legend - Modern renkli açıklamalar (parantez içinde toplam sayılar)
    legend_text = [
        f"● Total  ({total_sum})",
        f"○ Open  ({open_sum})", 
        f"◆ Review  ({review_sum})",
        f"■ Progress  ({progress_sum})",
        f"▲ Closed  ({closed_sum})"
    ]

    if language == 'tr':
        legend_text = [
            f"● Toplam  ({total_sum})",
            f"○ Acik  ({open_sum})", 
            f"◆ Inceleme  ({review_sum})",
            f"■ Ilerleme  ({progress_sum})",
            f"▲ Kapali  ({closed_sum})"
        ]
    
    legend_colors = [
        colors.blue,
        colors.red,
        colors.orange,
        colors.HexColor('#2196F3'),
        colors.green
    ]
    
    elements = []
    elements.append(Paragraph(title_text, title_style))
    elements.append(drawing)
    
    # Renkli legend ekle - her biri kendi renginde
    for i, text in enumerate(legend_text):
        legend_style = ParagraphStyle(
            'Legend',
            parent=getSampleStyleSheet()['Normal'],
            fontSize=10,
            spaceAfter=5,
            alignment=1,  # Ortalı
            textColor=legend_colors[i],
            fontName='Helvetica-Bold'
        )
        elements.append(Paragraph(text, legend_style))
    
    return elements
    
    elements = []
    elements.append(Paragraph(title_text, title_style))
    elements.append(drawing)
    
    # Legend ekle
    for text in legend_text:
        elements.append(Paragraph(text, legend_style))
    
    return elements

def create_daily_chat_line_plot(chats_data, language='tr'):
    """Daily chat distribution için line plot oluştur"""
    if not chats_data or not chats_data.get('dates'):
        return []
    
    dates_data = chats_data['dates']
    if not dates_data:
        return []
    
    # X ekseni için günleri hazırla (1'den başlayarak)
    x_values = list(range(1, len(dates_data) + 1))
    
    # Y ekseni için verileri hazırla
    chat_data = []
    messages_data = []
    supporter_data = []
    user_data = []
    
    for i, date_info in enumerate(dates_data):
        x_val = i + 1
        
        # Chat count
        chat_count = date_info.get('chatCount', 0)
        chat_data.append((x_val, chat_count))
        
        # Message count
        message_count = date_info.get('messageCount', 0)
        messages_data.append((x_val, message_count))
        
        # Supporter messages
        supporter_count = date_info.get('customerSupporterMessages', date_info.get('agentMessages', 0))
        supporter_data.append((x_val, supporter_count))
        
        # User messages
        user_count = date_info.get('userMessages', date_info.get('customerMessages', 0))
        user_data.append((x_val, user_count))
    
    # Line plot oluştur
    drawing = Drawing(500, 300)
    
    lp = LinePlot()
    lp.x = 80
    lp.y = 100  # Y pozisyonunu artır
    lp.height = 250  # Yüksekliği artır
    lp.width = 350
    lp.data = [chat_data, messages_data, supporter_data, user_data]
    lp.joinedLines = 1
    
    # Her çizgi için farklı marker ve renk
    lp.lines[0].symbol = makeMarker('FilledCircle')  # Chat - Mavi
    lp.lines[0].strokeColor = colors.blue
    lp.lines[0].strokeWidth = 2
    
    lp.lines[1].symbol = makeMarker('Circle')  # Messages - Yeşil
    lp.lines[1].strokeColor = colors.green
    lp.lines[1].strokeWidth = 2
    
    lp.lines[2].symbol = makeMarker('Diamond')  # Supporter - Turuncu
    lp.lines[2].strokeColor = colors.orange
    lp.lines[2].strokeWidth = 2
    
    lp.lines[3].symbol = makeMarker('Square')  # User - Kırmızı
    lp.lines[3].strokeColor = colors.red
    lp.lines[3].strokeWidth = 2
    
    lp.lineLabelFormat = '%2.0f'
    lp.strokeColor = colors.black
    
    # X ekseni ayarları
    lp.xValueAxis.valueMin = 0
    lp.xValueAxis.valueMax = len(dates_data) + 1
    lp.xValueAxis.valueSteps = x_values
    lp.xValueAxis.labelTextFormat = '%d'
    
    # Y ekseni ayarları - Dinamik değer (en yüksek + 10)
    max_value = max([
        max([y for x, y in chat_data]) if chat_data else 0,
        max([y for x, y in messages_data]) if messages_data else 0,
        max([y for x, y in supporter_data]) if supporter_data else 0,
        max([y for x, y in user_data]) if user_data else 0
    ])
    
    lp.yValueAxis.valueMin = 0
    lp.yValueAxis.valueMax = max_value + 10  # En yüksek değer + 10
    lp.yValueAxis.valueSteps = list(range(0, max_value + 11, max(1, (max_value + 10) // 5)))
    
    drawing.add(lp)
    
    # Başlık
    title_text = "Günlük Chat Dagilimi" if language == 'tr' else "Daily Chat Distribution"
    title_style = ParagraphStyle(
        'ChartTitle',
        parent=getSampleStyleSheet()['Heading3'],
        fontSize=14,
        textColor=colors.black,
        alignment=1,  # Ortalı
        spaceBefore=10,
        spaceAfter=80  # Daha fazla boşluk
    )
    
    # Toplam değerleri hesapla
    chat_sum = sum([y for x, y in chat_data]) if chat_data else 0
    messages_sum = sum([y for x, y in messages_data]) if messages_data else 0
    supporter_sum = sum([y for x, y in supporter_data]) if supporter_data else 0
    user_sum = sum([y for x, y in user_data]) if user_data else 0
    
    # Legend - Modern renkli açıklamalar (parantez içinde toplam sayılar
    legend_text = [
            f"● Sohbetler  ({chat_sum})",
            f"○ Mesajlar  ({messages_sum})", 
            f"◆ Destekci  ({supporter_sum})",
            f"■ Kullanici  ({user_sum})"
    ]
    
    if language == 'en':
        legend_text = [
            f"● Chats  ({chat_sum})",
            f"○ Messages  ({messages_sum})", 
            f"◆ Supporter  ({supporter_sum})",
            f"■ User  ({user_sum})"
        ] 
    
    legend_colors = [
        colors.blue,
        colors.green,
        colors.orange,
        colors.red
    ]
    
    elements = []
    elements.append(Paragraph(title_text, title_style))
    elements.append(drawing)
    
    # Renkli legend ekle - her biri kendi renginde
    for i, text in enumerate(legend_text):
        legend_style = ParagraphStyle(
            'Legend',
            parent=getSampleStyleSheet()['Normal'],
            fontSize=10,
            spaceAfter=5,
            alignment=1,  # Ortalı
            textColor=legend_colors[i],
            fontName='Helvetica-Bold'
        )
        elements.append(Paragraph(text, legend_style))
    
    return elements

def create_python_table(headers, data, title=None, is_daily_table=False):
    """Python tarzı tablo oluştur"""
    if not data:
        return Spacer(1, 10)
    
    # Tablo verilerini hazırla
    table_data = [headers]
    for row in data:
        table_data.append(row)
    
    # Dinamik sütun genişliklerini hesapla
    col_widths = []
    for col_idx in range(len(headers)):
        # Bu sütundaki tüm değerleri topla (başlık dahil)
        column_values = [headers[col_idx]]
        for row in data:
            if col_idx < len(row):
                column_values.append(str(row[col_idx]))
        
        # En uzun değeri bul
        max_width = max(len(str(val)) for val in column_values)
        
        # Sayı sütunu mu kontrol et
        is_number_column = False
        if col_idx >= 2:  # İlk iki sütun genellikle text
            # Bu sütundaki değerlerin çoğu sayı mı kontrol et
            number_count = 0
            total_count = 0
            
            for row in data:
                if col_idx < len(row):
                    total_count += 1
                    try:
                        # Sayıya çevrilebiliyor mu kontrol et
                        float(str(row[col_idx]).replace(',', ''))
                        number_count += 1
                    except (ValueError, TypeError):
                        pass
            
            # %80'den fazlası sayı ise sayı sütunu kabul et
            if total_count > 0 and (number_count / total_count) >= 0.8:
                is_number_column = True
        
        if is_number_column:
            # Sayı sütunları için sadece sayı genişliği + padding
            col_width = max_width * 6 + 20  # 6px per character + 20px padding
        else:
            # Text sütunları için en büyük text genişliği + padding
            col_width = max_width * 8 + 30  # 8px per character + 30px padding
        
        # Minimum ve maksimum genişlik sınırları
        if is_daily_table:
            col_width = max(35, min(col_width, 80))  # Daily tablolar için sınırlar
        else:
            col_width = max(60, min(col_width, 150))  # Normal tablolar için sınırlar
        
        col_widths.append(col_width)
    
    # Tablo oluştur
    table = Table(table_data, colWidths=col_widths, repeatRows=1, spaceBefore=6, spaceAfter=6)
    
    # Tablo stilini ayarla
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f8f9fa')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#424242')),
        # Başlık satırı hizalaması sütun bazında yapılacak
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9 if is_daily_table else 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12 if is_daily_table else 14),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e9ecef')),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8 if is_daily_table else 9),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#666666')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('KEEPWITHNEXT', (0, 0), (-1, -1), 1),
        ('LEFTPADDING', (0, 0), (-1, -1), 6 if is_daily_table else 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6 if is_daily_table else 10),
        ('TOPPADDING', (0, 0), (-1, -1), 5 if is_daily_table else 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 5 if is_daily_table else 8),
    ])
    
    # Sütun türüne göre hizalama
    for col_idx in range(len(headers)):
        # Sayı sütunu mu kontrol et
        is_number_column = False
        if col_idx >= 2:  # İlk iki sütun genellikle text
            # Bu sütundaki değerlerin çoğu sayı mı kontrol et
            number_count = 0
            total_count = 0
            
            for row in data:
                if col_idx < len(row):
                    total_count += 1
                    try:
                        # Sayıya çevrilebiliyor mu kontrol et
                        float(str(row[col_idx]).replace(',', ''))
                        number_count += 1
                    except (ValueError, TypeError):
                        pass
            
            # %80'den fazlası sayı ise sayı sütunu kabul et
            if total_count > 0 and (number_count / total_count) >= 0.8:
                is_number_column = True
        
        if is_number_column:
            # Sayı sütunları ortalı
            style.add('ALIGN', (col_idx, 1), (col_idx, -1), 'CENTER')
            # Sayı sütunları başlığı da ortalı
            style.add('ALIGN', (col_idx, 0), (col_idx, 0), 'CENTER')
        else:
            # Text sütunları sol hizalı
            style.add('ALIGN', (col_idx, 1), (col_idx, -1), 'LEFT')
            # Text sütunları başlığı da sol hizalı
            style.add('ALIGN', (col_idx, 0), (col_idx, 0), 'LEFT')
    
    # Metin sarma özelliği ekle
    style.add('WORDWRAP', (0, 0), (-1, -1), True)
    
    table.setStyle(style)
    
    # Başlık ve tabloyu birlikte tut
    elements = []
    if title:
        elements.append(KeepTogether([create_subsection_title(title), table]))
    else:
        elements.append(table)
    return elements

def create_task_statistics_tables(tasks_data, language='tr'):
    """Task istatistikleri için tablolar oluştur"""
    elements = []
    
    if not tasks_data:
        return elements
    
    # Status Distribution - Pie Chart
    if tasks_data.get('status'):
        elements.extend(create_status_pie_chart(tasks_data['status'], language))
        elements.append(Spacer(1, 15))
    
    # Priority Distribution - Pie Chart
    if tasks_data.get('priority'):
        elements.extend(create_priority_pie_chart(tasks_data['priority'], language))
        elements.append(Spacer(1, 15))
    
    # Leader Distribution
    if tasks_data.get('leader'):
        headers = [
            "Leader Name" if language == 'en' else "Lider Adi",
            "Email",
            "Assigned Tasks" if language == 'en' else "Atanan Task",
            "Completed Tasks" if language == 'en' else "Tamamlanan Task",
            "Overdue Tasks" if language == 'en' else "Geciken Task"
        ]
        data = [
            [
                str(l.get('name', l.get('id', ''))),
                str(l.get('email', 'N/A')),
                str(l.get('assignTaskCount', 0)),
                str(l.get('doneTaskCount', 0)),
                str(l.get('overDueTaskCount', 0))
            ]
            for l in tasks_data['leader']
        ]
        title = "Leader Task Distribution" if language == 'en' else "Lider Task Dagilimi"
        elements.extend(create_python_table(headers, data, title))
        elements.append(Spacer(1, 10))
    
    # Employee Distribution
    if tasks_data.get('employee'):
        headers = [
            "Employee Name" if language == 'en' else "Calisan Adi",
            "Email",
            "Completed Tasks" if language == 'en' else "Tamamlanan Task",
            "Overdue Tasks" if language == 'en' else "Geciken Task"
        ]
        data = [
            [
                str(e.get('name', e.get('id', ''))),
                str(e.get('email', 'N/A')),
                str(e.get('doneTaskCount', 0)),
                str(e.get('overDueTaskCount', 0))
            ]
            for e in tasks_data['employee']
        ]
        title = "Employee Task Distribution" if language == 'en' else "Calisan Task Dagilimi"
        elements.extend(create_python_table(headers, data, title))
        elements.append(Spacer(1, 10))
    
    return elements

def create_ticket_statistics_tables(tickets_data, language='tr'):
    """Ticket istatistikleri için tablolar oluştur"""
    elements = []
    
    if not tickets_data:
        return elements
    
    # Status Distribution - Pie Chart
    if tickets_data.get('status'):
        elements.extend(create_status_pie_chart(tickets_data['status'], language))
        elements.append(Spacer(1, 15))
    
    # Resolution Times - Pie Chart
    if tickets_data.get('resolveTimes'):
        elements.extend(create_resolution_pie_chart(tickets_data['resolveTimes'], language))
        elements.append(Spacer(1, 15))
    
    # Unassigned Tickets
    if tickets_data.get('numberOfNotAssignedAgentTickets') is not None:
        headers = ["Info" if language == 'en' else "Bilgi", "Value" if language == 'en' else "Değer"]
        data = [
            ["Unassigned Ticket Count" if language == 'en' else "Atanmamis Ticket Sayisi", 
             str(tickets_data['numberOfNotAssignedAgentTickets'])]
        ]
        title = "Unassigned Ticket Information" if language == 'en' else "Atanmamis Ticket Bilgisi"
        elements.extend(create_python_table(headers, data, title))
        elements.append(Spacer(1, 10))
    
    # Daily Distribution - Line Plot
    if tickets_data.get('dates'):
        elements.extend(create_daily_ticket_line_plot(tickets_data, language))
        elements.append(Spacer(1, 15))
    
    return elements

def create_user_statistics_tables(users_data, language='tr'):
    """User istatistikleri için tablolar oluştur"""
    elements = []
    
    if not users_data:
        return elements
    
    # General User Information - Pie Chart
    if users_data.get('total') is not None:
        elements.extend(create_user_verification_pie_chart(users_data, language))
        elements.append(Spacer(1, 15))
    
    # Role Distribution - Pie Chart
    if users_data.get('roles'):
        elements.extend(create_user_role_pie_chart(users_data['roles'], language))
        elements.append(Spacer(1, 15))
    
    return elements

def create_chat_statistics_tables(chats_data, language='tr'):
    """Chat istatistikleri için tablolar oluştur"""
    elements = []
    
    if not chats_data:
        return elements
    
    # General Chat Information
    if chats_data.get('chatCount') is not None or chats_data.get('messageCount') is not None:
        headers = ["Info" if language == 'en' else "Bilgi", "Value" if language == 'en' else "Değer"]
        data = [
            ["Total Chat Count" if language == 'en' else "Toplam Chat Sayisi", str(chats_data.get('chatCount', 0))],
            ["Total Message Count" if language == 'en' else "Toplam Mesaj Sayisi", str(chats_data.get('messageCount', 0))]
        ]
        elements.extend(create_python_table(headers, data, None))  # Başlık yok
        elements.append(Spacer(1, 10))
    
    # Daily Chat Distribution - Line Plot
    if chats_data.get('dates'):
        elements.extend(create_daily_chat_line_plot(chats_data, language))
        elements.append(Spacer(1, 15))
    
    return elements

def create_category_statistics_tables(categories_data, language='tr'):
    """Category istatistikleri için tablolar oluştur"""
    elements = []
    
    if not categories_data:
        return elements
    
    headers = [
        "Category Name" if language == 'en' else "Kategori Adi",
        "Leader Count" if language == 'en' else "Lider Sayisi",
        "Product Count" if language == 'en' else "Ürün Sayisi",
        "Related Ticket Count" if language == 'en' else "Iliskili Ticket Sayisi"
    ]
    
    data = [
        [
            str(c.get('categoryNameEn' if language == 'en' else 'categoryNameTr', 'N/A')),
            str(c.get('numberOfLeader', 0)),
            str(c.get('numberOfProduct', 0)),
            str(c.get('numberOfRelatedTicket', 0))
        ]
        for c in categories_data
    ]
    
    elements.extend(create_python_table(headers, data, None))  # Başlık yok
    elements.append(Spacer(1, 10))
    
    return elements

def create_product_statistics_tables(products_data, categories_data, language='tr'):
    """Product istatistikleri için tablolar oluştur"""
    elements = []
    
    if not products_data:
        return elements
    
    # Kategori ID'lerini kategori adlarına çevirmek için mapping oluştur
    category_mapping = {}
    if categories_data:
        for cat in categories_data:
            category_mapping[cat.get('id')] = cat.get('categoryNameEn' if language == 'en' else 'categoryNameTr')
    
    headers = [
        "Product Name" if language == 'en' else "Ürün Adi",
        "Category Name" if language == 'en' else "Kategori Adi",
        "Related Ticket Count" if language == 'en' else "İlişkili Ticket Sayisi"
    ]
    
    data = []
    for p in products_data:
        category_id = p.get('productCategoryId')
        category_name = category_mapping.get(category_id, str(category_id) if category_id else 'N/A')
        
        data.append([
            str(p.get('productNameEn' if language == 'en' else 'productNameTr', 'N/A')),
            str(category_name),
            str(p.get('relatedTicketCount', 0))
        ])
    
    elements.extend(create_python_table(headers, data, None))  # Başlık yok
    elements.append(Spacer(1, 10))
    
    return elements

def create_pdf_with_python_tables(export_data, language='tr'):
    """Python tabloları ile PDF oluştur"""
    try:
        logger.info(f"Python tabloları ile PDF oluşturma başlatılıyor... Language: {language}")
        
        # PDF dokümanı oluştur
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        
        # Stil şablonları
        styles = getSampleStyleSheet()
        
        # Başlık
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=20,
            textColor=colors.HexColor('#1976d2'),
            alignment=1  # Center
        )
        
        # Bölüm başlığı
        section_style = ParagraphStyle(
            'SectionTitle',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=10,
            textColor=colors.HexColor('#424242'),
            spaceBefore=20
        )
        
        # Tarih formatı
        generation_date = datetime.now().strftime("%d.%m.%Y %H:%M") if language == 'tr' else datetime.now().strftime("%m/%d/%Y %H:%M")
        
        # PDF içeriği
        story = []
        
        # Ana başlık
        title_text = "Panel Istatistikleri Raporu" if language == 'tr' else "Dashboard Statistics Report"
        story.append(Paragraph(title_text, title_style))
        
        # Tarih
        date_text = f"Tarih: {generation_date}" if language == 'tr' else f"Date: {generation_date}"
        story.append(Paragraph(date_text, styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Task İstatistikleri
        if export_data.get('tasks'):
            section_title = "Task Istatistikleri" if language == 'tr' else "Task Statistics"
            story.append(Paragraph(section_title, section_style))
            story.extend(create_task_statistics_tables(export_data['tasks'], language))
        
        # Ticket İstatistikleri
        if export_data.get('tickets'):
            story.append(PageBreak())  # Sayfa geçişi
            section_title = "Ticket Istatistikleri" if language == 'tr' else "Ticket Statistics"
            story.append(Paragraph(section_title, section_style))
            story.extend(create_ticket_statistics_tables(export_data['tickets'], language))
        
        # User İstatistikleri
        if export_data.get('users'):
            story.append(PageBreak())  # Sayfa geçişi
            section_title = "Kullanici Istatistikleri" if language == 'tr' else "User Statistics"
            story.append(Paragraph(section_title, section_style))
            story.extend(create_user_statistics_tables(export_data['users'], language))
        
        # Chat İstatistikleri
        if export_data.get('chats'):
            story.append(PageBreak())  # Sayfa geçişi
            section_title = "Chat Istatistikleri" if language == 'tr' else "Chat Statistics"
            story.append(Paragraph(section_title, section_style))
            story.extend(create_chat_statistics_tables(export_data['chats'], language))
        
        # Category İstatistikleri
        if export_data.get('categories'):
            story.append(PageBreak())  # Sayfa geçişi
            section_title = "Kategori Istatistikleri" if language == 'tr' else "Category Statistics"
            story.append(Paragraph(section_title, section_style))
            story.extend(create_category_statistics_tables(export_data['categories'], language))
        
        # Product İstatistikleri
        if export_data.get('products'):
            story.append(PageBreak())  # Sayfa geçişi
            section_title = "Ürün Istatistikleri" if language == 'tr' else "Product Statistics"
            story.append(Paragraph(section_title, section_style))
            story.extend(create_product_statistics_tables(export_data['products'], export_data.get('categories', []), language))
        
        # PDF oluştur
        doc.build(story)
        
        buffer.seek(0)
        pdf_bytes = buffer.getvalue()
        buffer.close()
        
        logger.info(f"Python tabloları ile PDF başarıyla oluşturuldu. Size: {len(pdf_bytes)} bytes")
        
        return base64.b64encode(pdf_bytes).decode('utf-8')
        
    except Exception as e:
        logger.error(f"Python tabloları ile PDF oluşturulamadi: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return None

def create_csv_content(export_data, language='tr'):
    """CSV icerigini olustur"""
    try:
        output = BytesIO()
        writer = csv.writer(output)
        
        # Baslik satiri
        if language == 'tr':
            writer.writerow(['Veri Tipi', 'Icerik'])
        else:
            writer.writerow(['Data Type', 'Content'])
        
        # Tasks
        if export_data.get('tasks'):
            if language == 'tr':
                writer.writerow(['Gorevler', ''])
            else:
                writer.writerow(['Tasks', ''])
            for task in export_data['tasks']:
                writer.writerow(['', str(task)])
            writer.writerow([])
        
        # Tickets
        if export_data.get('tickets'):
            if language == 'tr':
                writer.writerow(['Ticketlar', ''])
            else:
                writer.writerow(['Tickets', ''])
            for ticket in export_data['tickets']:
                writer.writerow(['', str(ticket)])
            writer.writerow([])
        
        # Users
        if export_data.get('users'):
            if language == 'tr':
                writer.writerow(['Kullanicilar', ''])
            else:
                writer.writerow(['Users', ''])
            for user in export_data['users']:
                writer.writerow(['', str(user)])
            writer.writerow([])
        
        output.seek(0)
        csv_content = output.getvalue().decode('utf-8')
        output.close()
        return csv_content
    except Exception as e:
        logger.error(f"CSV content olusturulamadi: {e}")
        if language == 'tr':
            return "Veri Tipi,Icerik\nHata,CSV olusturma basarisiz"
        else:
            return "Data Type,Content\nError,CSV generation failed"

def create_pdf_content(export_data, language='tr'):
    """PDF içeriğini oluştur"""
    try:
        if language == 'tr':
            pdf_content = """
Panel İstatistikleri Özeti

• Kullanıcılar: Toplam, doğrulanmış, engellenmiş, roller
• Aktif Temsilciler: Son 24 saat
• Ticketler: Toplam, açık, atanmamış, durumlar, kapanma süreleri
• Tasklar: Toplam, durumlar, öncelik, lider/çalışan dağılımı
• Ürünler: Toplam ürün
• Kategoriler: Kategori bazlı dağılım
• Chat/Mesaj: Toplam chat, mesaj, trendler
• Kullanıcı Kayıt Trendi

Detaylı veriler için lütfen PDF içeriğine bakınız.
            """
        else:
            pdf_content = """
Dashboard Statistics Summary

• Users: Total, verified, blocked, roles
• Active Representatives: Last 24 hours
• Tickets: Total, open, unassigned, statuses, resolution times
• Tasks: Total, statuses, priority, leader/employee distribution
• Products: Total products
• Categories: Category-based distribution
• Chat/Messages: Total chat, messages, trends
• User Registration Trend

Please check the PDF content for detailed data.
            """
        
        return pdf_content.strip()
    except Exception as e:
        logger.error(f"PDF content oluşturulamadi: {e}")
        return "Dashboard Statistics Summary"

def create_pdf_file(export_data, language='tr'):
    """PDF dosyasını oluştur ve base64'e çevir - Python tabloları ile"""
    try:
        logger.info(f"PDF oluşturma başlatılıyor... Language: {language}")
        
        # Python tabloları ile PDF oluştur
        return create_pdf_with_python_tables(export_data, language)
        
    except Exception as e:
        logger.error(f"PDF dosyası oluşturulamadi: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return None

def create_section_title(title):
    """Bölüm başlığı oluştur"""
    style = ParagraphStyle(
        'SectionTitle',
        parent=getSampleStyleSheet()['Heading2'],
        fontSize=16,
        spaceAfter=15,
        textColor=colors.HexColor('#1976d2'),
        fontName='Helvetica-Bold',
        leftIndent=0,
        backColor=colors.HexColor('#f5f5f5'),
        keepWithNext=1  # Sonraki içerikle birlikte tut
    )
    return Paragraph(title, style)

def create_subsection_title(title):
    """Alt bölüm başlığı oluştur"""
    style = ParagraphStyle(
        'SubsectionTitle',
        parent=getSampleStyleSheet()['Normal'],
        fontSize=12,
        spaceAfter=8,
        textColor=colors.HexColor('#424242'),
        fontName='Helvetica-Bold',
        alignment=1,  # Ortalı
        keepWithNext=1  # Sonraki içerikle birlikte tut
    )
    return Paragraph(title, style)

def create_data_table(data_dict):
    """Veri tablosu oluştur"""
    if not data_dict:
        return Spacer(1, 10)

    # Tablo verilerini hazırla
    table_data = [['Label', 'Value']]
    for key, value in data_dict.items():
        table_data.append([str(key), str(value)])

    # Tablo oluştur
    table = Table(table_data, colWidths=[200, 100])

    # Tablo stilini ayarla
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f8f9fa')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#424242')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e9ecef')),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#666666')),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
        ('TEXTCOLOR', (1, 1), (1, -1), colors.HexColor('#1976d2')),
        ('FONTNAME', (1, 1), (1, -1), 'Helvetica-Bold'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('KEEPWITHNEXT', (0, 0), (-1, -1), 1),  # Tablo satırlarını birlikte tut
    ])

    table.setStyle(style)
    
    # Tablo için özel stil oluştur - sayfa sonunda kesilmesini önle
    table_style = ParagraphStyle(
        'TableStyle',
        parent=getSampleStyleSheet()['Normal'],
        pageBreakBefore=0,  # Sayfa sonunda kesilmesini önle
        keepWithNext=1,     # Sonraki içerikle birlikte tut
        spaceAfter=10
    )
    
    # Tabloyu bir container içine al
    from reportlab.platypus import KeepTogether
    return KeepTogether([table])

def write_section_title(ws, row, title):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(size=14, bold=True)
    cell.alignment = Alignment(horizontal='left', vertical='center')

def write_table(ws, start_row, headers, rows):
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = Font(bold=True)
    for row_index, row_data in enumerate(rows, start=start_row+1):
        for col_index, cell_value in enumerate(row_data, start=1):
            # None değerleri boş string olarak göster
            if cell_value is None:
                cell_value = ''
            ws.cell(row=row_index, column=col_index, value=cell_value)
    return start_row + len(rows) + 2  # 2 satır boşluk bırak

def create_excel_file(export_data, language='tr'):
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard Report"
    current_row = 1

    # Baslik ve tarih
    if language == 'tr':
        ws.cell(row=current_row, column=1, value="Dashboard Istatistik Raporu").font = Font(size=16, bold=True)
        current_row += 1
        ws.cell(row=current_row, column=1, value=f"Tarih: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    else:
        ws.cell(row=current_row, column=1, value="Dashboard Statistics Report").font = Font(size=16, bold=True)
        current_row += 1
        ws.cell(row=current_row, column=1, value=f"Date: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    current_row += 2

    # TASK STATISTICS
    tasks = export_data.get('tasks', {})
    if tasks:
        if 'status' in tasks:
            if language == 'tr':
                write_section_title(ws, current_row, "Gorev Istatistikleri - Durum Dagilimi")
                current_row += 1
                headers = ["Durum", "Sayi"]
            else:
                write_section_title(ws, current_row, "Task Statistics - Status Distribution")
                current_row += 1
                headers = ["Status", "Count"]
            rows = [[k, v] for k, v in tasks['status'].items()]
            current_row = write_table(ws, current_row, headers, rows)
        if 'priority' in tasks:
            if language == 'tr':
                write_section_title(ws, current_row, "Gorev Istatistikleri - Oncelik Dagilimi")
                current_row += 1
                headers = ["Oncelik", "Sayi"]
            else:
                write_section_title(ws, current_row, "Task Statistics - Priority Distribution")
                current_row += 1
                headers = ["Priority", "Count"]
            rows = [[k, v] for k, v in tasks['priority'].items()]
            current_row = write_table(ws, current_row, headers, rows)
        if 'leader' in tasks:
            if language == 'tr':
                write_section_title(ws, current_row, "Lider Gorev Dagilimi")
                current_row += 1
                headers = ["Lider Adi", "E-posta", "Atanan Gorevler", "Tamamlanan", "Geciken"]
            else:
                write_section_title(ws, current_row, "Leader Task Distribution")
                current_row += 1
                headers = ["Leader Name", "Email", "Assigned Tasks", "Completed", "Overdue"]
            rows = [
                [
                    l.get('name', ''),
                    l.get('email', ''),
                    l.get('assignTaskCount', ''),
                    l.get('doneTaskCount', ''),
                    l.get('overDueTaskCount', '')
                ]
                for l in tasks['leader']
            ]
            current_row = write_table(ws, current_row, headers, rows)
        if 'employee' in tasks:
            if language == 'tr':
                write_section_title(ws, current_row, "Calisan Gorev Dagilimi")
                current_row += 1
                headers = ["Calisan Adi", "E-posta", "Tamamlanan", "Geciken"]
            else:
                write_section_title(ws, current_row, "Employee Task Distribution")
                current_row += 1
                headers = ["Employee Name", "Email", "Completed", "Overdue"]
            rows = [
                [
                    e.get('name', ''),
                    e.get('email', ''),
                    e.get('doneTaskCount', ''),
                    e.get('overDueTaskCount', '')
                ]
                for e in tasks['employee']
            ]
            current_row = write_table(ws, current_row, headers, rows)

    # TICKET STATISTICS
    tickets = export_data.get('tickets', {})
    if tickets:
        if 'status' in tickets:
            if language == 'tr':
                write_section_title(ws, current_row, "Ticket Istatistikleri - Durum")
                current_row += 1
                headers = ["Durum", "Sayi"]
            else:
                write_section_title(ws, current_row, "Ticket Statistics - Status")
                current_row += 1
                headers = ["Status", "Count"]
            rows = [[k, v] for k, v in tickets['status'].items()]
            current_row = write_table(ws, current_row, headers, rows)
        if 'resolveTimes' in tickets:
            if language == 'tr':
                write_section_title(ws, current_row, "Cozum Sureleri")
                current_row += 1
                headers = ["Sure Araligi", "Sayi"]
            else:
                write_section_title(ws, current_row, "Resolution Times")
                current_row += 1
                headers = ["Time Range", "Count"]
            rows = [[k, v] for k, v in tickets['resolveTimes'].items()]
            current_row = write_table(ws, current_row, headers, rows)
        if 'numberOfNotAssignedAgentTickets' in tickets:
            if language == 'tr':
                write_section_title(ws, current_row, "Atanmamis Ticket Bilgisi")
                current_row += 1
                headers = ["Bilgi", "Deger"]
                rows = [["Atanmamis Ticket Sayisi", tickets['numberOfNotAssignedAgentTickets']]]
            else:
                write_section_title(ws, current_row, "Unassigned Ticket Information")
                current_row += 1
                headers = ["Info", "Value"]
                rows = [["Unassigned Ticket Count", tickets['numberOfNotAssignedAgentTickets']]]
            current_row = write_table(ws, current_row, headers, rows)
        if 'dates' in tickets:
            if language == 'tr':
                write_section_title(ws, current_row, "Gunluk Ticket Dagilimi")
                current_row += 1
                headers = ["Tarih", "Gun", "Toplam", "Acik", "Incelemede", "Devam Eden", "Kapali"]
            else:
                write_section_title(ws, current_row, "Daily Ticket Distribution")
                current_row += 1
                headers = ["Date", "Day", "Total", "Open", "In Review", "In Progress", "Closed"]
            rows = [[d.get('date', ''), d.get('dayName', ''), d.get('totalTickets', ''), d.get('openTickets', ''), d.get('inReviewTickets', ''), d.get('inProgressTickets', ''), d.get('closedTickets', '')] for d in tickets['dates']]
            current_row = write_table(ws, current_row, headers, rows)

    # USER STATISTICS
    users = export_data.get('users', {})
    if users:
        if language == 'tr':
            write_section_title(ws, current_row, "Kullanici Istatistikleri - Genel Bilgi")
            current_row += 1
            headers = ["Bilgi", "Deger"]
            rows = [
                ["Toplam Kullanici", users.get('total')],
                ["Dogrulanmis Kullanici", users.get('verifiedUsers')],
                ["Engellenmis Kullanici", users.get('blockedUsers')]
            ]
        else:
            write_section_title(ws, current_row, "User Statistics - General Info")
            current_row += 1
            headers = ["Info", "Value"]
            rows = [
                ["Total Users", users.get('total')],
                ["Verified Users", users.get('verifiedUsers')],
                ["Blocked Users", users.get('blockedUsers')]
            ]
        current_row = write_table(ws, current_row, headers, rows)
        if 'roles' in users:
            if language == 'tr':
                write_section_title(ws, current_row, "Rol Bazli Kullanici Dagilimi")
                current_row += 1
                headers = ["Rol", "Sayi"]
            else:
                write_section_title(ws, current_row, "User Distribution by Role")
                current_row += 1
                headers = ["Role", "Count"]
            rows = [[r.get('roleName'), r.get('count', r.get('numberOfUsers'))] for r in users['roles']]
            current_row = write_table(ws, current_row, headers, rows)

    # CHATS
    chats = export_data.get('chats', {})
    if chats:
        if language == 'tr':
            write_section_title(ws, current_row, "Genel Sohbet Bilgisi")
            current_row += 1
            headers = ["Bilgi", "Deger"]
            rows = [
                ["Toplam Sohbet", chats.get('chatCount')],
                ["Toplam Mesaj", chats.get('messageCount')]
            ]
        else:
            write_section_title(ws, current_row, "General Chat Information")
            current_row += 1
            headers = ["Info", "Value"]
            rows = [
                ["Total Chats", chats.get('chatCount')],
                ["Total Messages", chats.get('messageCount')]
            ]
        current_row = write_table(ws, current_row, headers, rows)
        if 'dates' in chats:
            if language == 'tr':
                write_section_title(ws, current_row, "Gunluk Sohbet Dagilimi")
                current_row += 1
                headers = ["Tarih", "Gun", "Sohbet Sayisi", "Toplam Mesaj", "Destek Mesaji", "Kullanici Mesaji"]
            else:
                write_section_title(ws, current_row, "Daily Chat Distribution")
                current_row += 1
                headers = ["Date", "Day", "Chat Count", "Total Messages", "Supporter Msg", "User Msg"]
            rows = [[d.get('date', ''), d.get('dayName', ''), d.get('chatCount'), d.get('messageCount'), d.get('agentMessages'), d.get('customerMessages')] for d in chats['dates']]
            current_row = write_table(ws, current_row, headers, rows)

    # CATEGORIES
    categories = export_data.get('categories', [])
    if categories:
        if language == 'tr':
            write_section_title(ws, current_row, "Kategori Istatistikleri")
        else:
            write_section_title(ws, current_row, "Category Statistics")
        current_row += 1
        if language == 'tr':
            headers = ["Kategori Adi", "Lider Sayisi", "Urun Sayisi", "Iliskili Ticket Sayisi"]
        else:
            headers = ["Category Name", "Leader Count", "Product Count", "Related Ticket Count"]
        rows = [[c.get('categoryNameTr') or c.get('categoryNameEn'), c.get('numberOfLeader'), c.get('numberOfProduct'), c.get('numberOfRelatedTicket')] for c in categories]
        current_row = write_table(ws, current_row, headers, rows)

    # PRODUCTS
    products = export_data.get('products', [])
    if products:
        if language == 'tr':
            write_section_title(ws, current_row, "Urun Istatistikleri")
            current_row += 1
            headers = ["Urun Adi", "Kategori Adi", "Iliskili Ticket Sayisi"]
        else:
            write_section_title(ws, current_row, "Product Statistics")
            current_row += 1
            headers = ["Product Name", "Category Name", "Related Ticket Count"]
        
        # Kategori ID'lerini kategori adlarına cevirmek icin mapping olustur
        categories = export_data.get('categories', [])
        category_mapping = {}
        for cat in categories:
            category_mapping[cat.get('id')] = cat.get('categoryNameTr') or cat.get('categoryNameEn')
        
        rows = []
        for p in products:
            category_id = p.get('productCategoryId')
            category_name = category_mapping.get(category_id, category_id)  # Eger mapping yoksa ID'yi goster
            rows.append([
                p.get('productNameTr') or p.get('productNameEn'),
                category_name,
                p.get('relatedTicketCount')
            ])
        current_row = write_table(ws, current_row, headers, rows)

    # Otomatik sutun genisligi
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Dosyayi BytesIO ile kaydet ve base64 dondur
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return base64.b64encode(output.read()).decode('utf-8')

def create_file_content(export_data, file_type, language='tr'):
    """Dosya turune gore icerik olustur"""
    if file_type == 'json':
        return base64.b64encode(json.dumps(export_data, indent=2, ensure_ascii=False).encode('utf-8')).decode('utf-8')
    elif file_type == 'csv':
        csv_content = create_csv_content(export_data, language)
        return base64.b64encode(csv_content.encode('utf-8')).decode('utf-8')
    elif file_type == 'pdf':
        return create_pdf_file(export_data, language)
    elif file_type == 'excel':
        return create_excel_file(export_data, language)
    else:
        # Default JSON
        return base64.b64encode(json.dumps(export_data, indent=2, ensure_ascii=False).encode('utf-8')).decode('utf-8')

def read_mail_template(language='tr'):
    """Mail template'ini oku"""
    try:
        template_name = 'dashboard_statictics_mail_content_tr.html' if language == 'tr' else 'dashboard_statictics_mail_content_en.html'
        template_path = f"templates/{template_name}"
        with open(template_path, 'r', encoding='utf-8') as f:
            return f.read()
    except Exception as e:
        logger.error(f"Mail template okunamadi: {e}")
        return "Dashboard export dosyaniz ektedir." if language == 'tr' else "Your dashboard export is attached." 